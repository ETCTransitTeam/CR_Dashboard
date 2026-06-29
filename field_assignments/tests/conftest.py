from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from field_assignments.core.constants import EXPECTED_COLUMNS

ROOT = Path(__file__).resolve().parents[2]
CLIENT_REFS_TEMPLATE = ROOT / "_client_refs" / "Runcut Template.xlsx"
SURVEY_TEMPLATE = ROOT / "Survey_Assignments_Files" / "Runcut Template.xlsx"


def _write_workbook(rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "RunCut"
    for column_index, header in enumerate(EXPECTED_COLUMNS, start=1):
        sheet.cell(row=1, column=column_index, value=header)
    for row_index, row_values in enumerate(rows, start=2):
        for column_index, value in enumerate(row_values, start=1):
            sheet.cell(row=row_index, column=column_index, value=value)
    out = BytesIO()
    workbook.save(out)
    return out.getvalue()


def _read_asn_values(source: bytes) -> list[str]:
    workbook = load_workbook(BytesIO(source), data_only=True)
    sheet = workbook.active
    values: list[str] = []
    for row_number in range(2, sheet.max_row + 1):
        value = sheet.cell(row=row_number, column=1).value
        values.append("" if value is None else str(value))
    return values


@pytest.fixture
def interline_workbook() -> bytes:
    """Block 10193 with Route 19 assignment rows and Route 10 interlined rows."""
    return _write_workbook(
        [
            [11, 10193, 19, "South", "5:30 AM", "Marion Transit Center", "Idaho st. @ O'Brien St.", "6:32 AM"],
            [11, 10193, 19, "North", "6:42 AM", "Idaho st. @ O'Brien St.", "Marion Transit Center", "7:45 AM"],
            ["", 10193, 10, "East", "8:00 AM", "Marion Transit Center", "Other Stop", "8:30 AM"],
            ["", 10193, 10, "West", "8:40 AM", "Other Stop", "Marion Transit Center", "9:10 AM"],
            [12, 10194, 30, "North", "7:00 AM", "Netpark Transfer Center", "End Stop", "8:00 AM"],
        ]
    )


@pytest.fixture
def blank_interline_workbook() -> bytes:
    """Blank Asn# rows on one block with mixed routes for fill tests."""
    return _write_workbook(
        [
            ["", 10193, 19, "South", "5:30 AM", "Marion Transit Center", "Idaho st. @ O'Brien St.", "6:32 AM"],
            ["", 10193, 10, "East", "8:00 AM", "Marion Transit Center", "Other Stop", "8:30 AM"],
            ["", 10193, 19, "North", "6:42 AM", "Idaho st. @ O'Brien St.", "Marion Transit Center", "7:45 AM"],
        ]
    )


@pytest.fixture
def start_location_workbook() -> bytes:
    return _write_workbook(
        [
            ["", 1, "1", "North", "6:00 AM", "University Area Transit Center", "Stop A", "6:30 AM"],
            ["", 2, "5", "South", "6:05 AM", "University Area Transit Center", "Stop B", "6:35 AM"],
            ["", 3, "6", "East", "6:10 AM", "Marion Transit Center", "Stop C", "6:40 AM"],
            ["", 4, "7", "West", "6:15 AM", "Marion Transit Center", "Stop D", "6:45 AM"],
            ["", 5, "8", "North", "6:20 AM", "Netpark Transfer Center", "Stop E", "6:50 AM"],
        ]
    )
