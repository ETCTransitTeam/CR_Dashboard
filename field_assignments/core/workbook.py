from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import BinaryIO

from openpyxl import Workbook, load_workbook

from field_assignments.core.constants import (
    COL_ASN,
    COL_BLOCK,
    COL_END_LOCATION,
    COL_ROUTE,
    COL_START_LOCATION,
    EXPECTED_COLUMNS,
)
from field_assignments.core.time_utils import is_blank, normalize_cell, normalize_header, option_values


def build_header_template() -> bytes:
    """Return a blank .xlsx workbook containing only the expected RunCut headers."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "RunCut"
    for column_index, header in enumerate(EXPECTED_COLUMNS, start=1):
        sheet.cell(row=1, column=column_index, value=header)
    out = BytesIO()
    workbook.save(out)
    return out.getvalue()


def read_headers(sheet) -> list[str]:
    raw = [normalize_header(sheet.cell(row=1, column=i).value) for i in range(1, 9)]
    return raw


def validate_headers(headers: list[str]) -> None:
    if headers != EXPECTED_COLUMNS:
        raise ValueError(
            "Workbook does not match the expected RunCut layout. "
            f"Expected {EXPECTED_COLUMNS}; found {headers}."
        )


def open_workbook(source: bytes | Path | BinaryIO, sheet_name: str | None = None, data_only: bool = True):
    if isinstance(source, Path):
        workbook = load_workbook(source, data_only=data_only)
    elif isinstance(source, bytes):
        workbook = load_workbook(BytesIO(source), data_only=data_only)
    else:
        workbook = load_workbook(source, data_only=data_only)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    headers = read_headers(sheet)
    validate_headers(headers)
    return workbook, sheet


def _route_location_maps(sheet) -> tuple[dict[str, list[str]], dict[str, list[str]]]:
    route_start: dict[str, set[str]] = {}
    route_end: dict[str, set[str]] = {}
    for row_number in range(2, sheet.max_row + 1):
        route = normalize_cell(sheet.cell(row=row_number, column=COL_ROUTE).value)
        if not route:
            continue
        start_loc = normalize_cell(sheet.cell(row=row_number, column=COL_START_LOCATION).value)
        end_loc = normalize_cell(sheet.cell(row=row_number, column=COL_END_LOCATION).value)
        if start_loc:
            route_start.setdefault(route, set()).add(start_loc)
        if end_loc:
            route_end.setdefault(route, set()).add(end_loc)
    start_map = {route: sorted(values) for route, values in route_start.items()}
    end_map = {route: sorted(values) for route, values in route_end.items()}
    return start_map, end_map


def workbook_options(source: bytes | Path | BinaryIO, sheet_name: str | None = None) -> dict[str, object]:
    workbook, sheet = open_workbook(source, sheet_name=sheet_name, data_only=True)

    blocks: list[object] = []
    routes: list[object] = []
    start_locations: list[object] = []
    end_locations: list[object] = []
    blank_rows = 0
    max_assignment = 0

    for row_number in range(2, sheet.max_row + 1):
        asn = sheet.cell(row=row_number, column=COL_ASN).value
        if is_blank(asn):
            blank_rows += 1
        else:
            try:
                max_assignment = max(max_assignment, int(float(normalize_cell(asn))))
            except ValueError:
                pass
        blocks.append(sheet.cell(row=row_number, column=COL_BLOCK).value)
        routes.append(sheet.cell(row=row_number, column=COL_ROUTE).value)
        start_locations.append(sheet.cell(row=row_number, column=COL_START_LOCATION).value)
        end_locations.append(sheet.cell(row=row_number, column=COL_END_LOCATION).value)

    route_start_locs, route_end_locs = _route_location_maps(sheet)

    return {
        "sheet": sheet.title,
        "blocks": option_values(blocks),
        "routes": option_values(routes),
        "start_locations": option_values(start_locations),
        "end_locations": option_values(end_locations),
        "route_start_locations": route_start_locs,
        "route_end_locations": route_end_locs,
        "blank_rows": blank_rows,
        "max_assignment": max_assignment,
        "next_assignment": max_assignment + 1,
    }
