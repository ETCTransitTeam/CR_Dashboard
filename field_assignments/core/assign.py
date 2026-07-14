from __future__ import annotations

import random
from io import BytesIO

from openpyxl import load_workbook

from field_assignments.core.constants import (
    ANY_LOCATION,
    ANY_ROUTE,
    COL_ASN,
    COL_BLOCK,
    COL_END_LOCATION,
    COL_END_TIME,
    COL_ROUTE,
    COL_START_LOCATION,
    COL_START_TIME,
    ROUTE_ONLY_MAX_GAP_MINUTES,
)
from field_assignments.core.time_utils import is_blank, normalize_cell, time_in_range, time_to_minutes
from field_assignments.core.workbook import workbook_options


def _is_any(value: object) -> bool:
    text = normalize_cell(value)
    return not text or text in {ANY_ROUTE, ANY_LOCATION, "(Any)", "Any"}


def _is_route_only(rules: dict[str, str]) -> bool:
    mode = str(rules.get("mode", "block") or "block").lower()
    if mode == "route_only":
        return True
    return str(rules.get("route_only", "false")).lower() in {"1", "true", "yes", "on"}


def _include_interlined(rules: dict[str, str]) -> bool:
    if _is_route_only(rules):
        return False
    return str(rules.get("include_interlined", "false")).lower() in {"1", "true", "yes", "on"}


def _requested_count(rules: dict[str, str]) -> int:
    try:
        return max(1, int(float(str(rules.get("count", "1") or "1"))))
    except ValueError:
        return 1


def row_matches_base_rules(
    sheet,
    row_number: int,
    rules: dict[str, str],
    block: str | None = None,
    *,
    require_route: bool = True,
    route_override: str | None = None,
) -> bool:
    if not is_blank(sheet.cell(row=row_number, column=COL_ASN).value):
        return False
    required_block = block if block is not None else rules.get("block", "")
    if required_block and normalize_cell(sheet.cell(row=row_number, column=COL_BLOCK).value) != required_block:
        return False
    required_route = route_override if route_override is not None else rules.get("route", "")
    if require_route and not _is_any(required_route):
        if normalize_cell(sheet.cell(row=row_number, column=COL_ROUTE).value) != normalize_cell(required_route):
            return False
    return True


def row_matches_segment_row(
    sheet,
    row_number: int,
    rules: dict[str, str],
    block: str,
    *,
    route_override: str | None = None,
) -> bool:
    """Match blank Asn# rows within an assignment segment."""
    if _is_route_only(rules):
        return row_matches_base_rules(
            sheet,
            row_number,
            rules,
            block=None,
            require_route=True,
            route_override=route_override or rules.get("route"),
        )
    return row_matches_base_rules(
        sheet,
        row_number,
        rules,
        block=block,
        require_route=not _include_interlined(rules),
        route_override=route_override,
    )


def row_matches_start_rules(
    sheet,
    row_number: int,
    rules: dict[str, str],
    *,
    route_override: str | None = None,
) -> bool:
    if not row_matches_base_rules(
        sheet,
        row_number,
        rules,
        require_route=not _is_any(route_override or rules.get("route")),
        route_override=route_override,
    ):
        return False
    start_location = rules.get("start_location", "")
    if not _is_any(start_location):
        if normalize_cell(sheet.cell(row=row_number, column=COL_START_LOCATION).value) != normalize_cell(
            start_location
        ):
            return False
    tolerance = int(rules.get("tolerance", "0") or 0)
    if not time_in_range(
        sheet.cell(row=row_number, column=COL_START_TIME).value,
        rules["start_from"],
        rules["start_to"],
        tolerance,
    ):
        return False
    return True


def row_matches_end_rules(
    sheet,
    row_number: int,
    rules: dict[str, str],
    block: str | None,
    *,
    route_override: str | None = None,
    require_block: bool = True,
) -> bool:
    if not row_matches_base_rules(
        sheet,
        row_number,
        rules,
        block=block if require_block else None,
        require_route=not _is_any(route_override or rules.get("route")),
        route_override=route_override,
    ):
        return False
    end_location = rules.get("end_location", "")
    if not _is_any(end_location):
        if normalize_cell(sheet.cell(row=row_number, column=COL_END_LOCATION).value) != normalize_cell(
            end_location
        ):
            return False
    tolerance = int(rules.get("tolerance", "0") or 0)
    if not time_in_range(
        sheet.cell(row=row_number, column=COL_END_TIME).value,
        rules["shift_from"],
        rules["shift_to"],
        tolerance,
    ):
        return False
    return True


def _iter_start_rows(sheet, scan_order: str) -> list[int]:
    rows = list(range(2, sheet.max_row + 1))
    if scan_order == "bottom_to_top":
        rows.reverse()
    elif scan_order == "random":
        random.shuffle(rows)
    return rows


def find_assignment_segment(
    sheet,
    rules: dict[str, str],
    *,
    route_override: str | None = None,
) -> tuple[int, int, str]:
    """Find a same-block assignment segment (default / interlined mode)."""
    for start_row in _iter_start_rows(sheet, rules.get("scan_order", "top_to_bottom")):
        if not row_matches_start_rules(sheet, start_row, rules, route_override=route_override):
            continue
        block = normalize_cell(sheet.cell(row=start_row, column=COL_BLOCK).value)
        for end_row in range(start_row, sheet.max_row + 1):
            if row_matches_end_rules(
                sheet,
                end_row,
                rules,
                block,
                route_override=route_override,
                require_block=True,
            ):
                return start_row, end_row, block
    raise ValueError(
        "No blank assignment segment matched those rules. "
        "Start Location and Start Time find the first row; End Location and Shift range find the final row."
    )


def _blank_route_row(
    sheet,
    row_number: int,
    route: str,
) -> bool:
    if not is_blank(sheet.cell(row=row_number, column=COL_ASN).value):
        return False
    if not _is_any(route) and normalize_cell(sheet.cell(row=row_number, column=COL_ROUTE).value) != normalize_cell(
        route
    ):
        return False
    return True


def _gap_minutes(prev_end: object, next_start: object) -> int | None:
    prev_min = time_to_minutes(prev_end)
    next_min = time_to_minutes(next_start)
    if prev_min is None or next_min is None:
        return None
    gap = next_min - prev_min
    if gap < 0:
        gap += 24 * 60
    return gap


def find_route_only_chain(
    sheet,
    rules: dict[str, str],
    *,
    route_override: str | None = None,
) -> tuple[int, int, str, list[int]]:
    """
    Chain trips on the same route across blocks.

    Continues while the next blank trip:
    - uses the same route
    - starts where the previous trip ended
    - boards within ROUTE_ONLY_MAX_GAP_MINUTES after previous alighting

    Ends early (hidden one-directional early-end) when no next trip is within the gap,
    even if the requested end window is later.
    """
    route = route_override if route_override is not None else rules.get("route", "")
    max_gap = int(rules.get("route_only_max_gap", ROUTE_ONLY_MAX_GAP_MINUTES) or ROUTE_ONLY_MAX_GAP_MINUTES)

    for start_row in _iter_start_rows(sheet, rules.get("scan_order", "top_to_bottom")):
        if not row_matches_start_rules(sheet, start_row, rules, route_override=route):
            continue

        chain = [start_row]
        used = {start_row}
        current = start_row
        route_value = normalize_cell(sheet.cell(row=start_row, column=COL_ROUTE).value)
        if not route_value:
            continue
        if not _is_any(route) and route_value != normalize_cell(route):
            continue

        while True:
            if row_matches_end_rules(
                sheet,
                current,
                rules,
                block=None,
                route_override=route_value,
                require_block=False,
            ):
                break

            prev_end_loc = normalize_cell(sheet.cell(row=current, column=COL_END_LOCATION).value)
            prev_end_time = sheet.cell(row=current, column=COL_END_TIME).value
            next_row: int | None = None

            for candidate in range(current + 1, sheet.max_row + 1):
                if candidate in used:
                    continue
                if not _blank_route_row(sheet, candidate, route_value):
                    continue
                start_loc = normalize_cell(sheet.cell(row=candidate, column=COL_START_LOCATION).value)
                if start_loc != prev_end_loc:
                    continue
                gap = _gap_minutes(prev_end_time, sheet.cell(row=candidate, column=COL_START_TIME).value)
                if gap is None or gap > max_gap:
                    continue
                next_row = candidate
                break

            if next_row is None:
                # Early end: no continuous trip within the max gap.
                break

            chain.append(next_row)
            used.add(next_row)
            current = next_row

        if len(chain) < 1:
            continue

        end_row = chain[-1]
        # Prefer chains that satisfy end location/time when specified; otherwise accept early end.
        end_ok = row_matches_end_rules(
            sheet,
            end_row,
            rules,
            block=None,
            route_override=route_value,
            require_block=False,
        )
        if not _is_any(rules.get("end_location", "")) and not end_ok:
            # If end location was required and we never reached it, still accept the early-end chain
            # only when we stopped due to a >60-min gap (surveyor would get a second assignment).
            pass

        block = normalize_cell(sheet.cell(row=start_row, column=COL_BLOCK).value)
        return start_row, end_row, block, chain

    raise ValueError(
        "No blank route-only assignment chain matched those rules. "
        "Trips must stay on the same route, connect end-to-start at the same location, "
        f"and board within {max_gap} minutes of the previous alighting."
    )


def _eligible_routes_for_any(sheet, rules: dict[str, str], options: dict[str, object]) -> list[str]:
    route_start = options.get("route_start_locations") or {}
    start_location = rules.get("start_location", "")
    routes = list(options.get("routes") or [])
    if _is_any(start_location):
        return routes

    matched: list[str] = []
    for route, locations in route_start.items():
        if normalize_cell(start_location) in {normalize_cell(loc) for loc in locations}:
            matched.append(str(route))
    if matched:
        return matched

    # Fallback: scan sheet for routes that actually start at the location.
    found: list[str] = []
    seen: set[str] = set()
    for row_number in range(2, sheet.max_row + 1):
        if normalize_cell(sheet.cell(row=row_number, column=COL_START_LOCATION).value) != normalize_cell(
            start_location
        ):
            continue
        route = normalize_cell(sheet.cell(row=row_number, column=COL_ROUTE).value)
        if route and route not in seen:
            seen.add(route)
            found.append(route)
    return found or routes


def _find_one_assignment(
    sheet,
    rules: dict[str, str],
    *,
    route_override: str | None = None,
) -> tuple[int, int, str, list[int]]:
    if _is_route_only(rules):
        return find_route_only_chain(sheet, rules, route_override=route_override)

    start_row, end_row, block = find_assignment_segment(sheet, rules, route_override=route_override)
    matching_rows: list[int] = []
    for row_number in range(start_row, end_row + 1):
        if row_matches_segment_row(sheet, row_number, rules, block, route_override=route_override):
            matching_rows.append(row_number)
    if not matching_rows:
        raise ValueError("A segment was found, but no blank Asn# rows were available to fill.")
    return start_row, end_row, block, matching_rows


def _apply_assignment(
    sheet,
    assignment_number: int,
    matching_rows: list[int],
) -> None:
    for row_number in matching_rows:
        sheet.cell(row=row_number, column=COL_ASN).value = assignment_number


def fill_assignment_numbers(
    source: bytes,
    sheet_name: str | None,
    rules_list: list[dict[str, str]],
) -> tuple[bytes, list[dict[str, object]]]:
    workbook = load_workbook(BytesIO(source))
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    options = workbook_options(source, sheet.title)
    next_assignment = int(options["next_assignment"])
    results: list[dict[str, object]] = []

    for index, rules in enumerate(rules_list, start=1):
        wanted = _requested_count(rules)
        created = 0
        any_route = _is_any(rules.get("route", ""))

        if any_route:
            # Round-robin: up to 2 assignments per route per pass, then cycle.
            routes = _eligible_routes_for_any(sheet, rules, options)
            if not routes:
                raise ValueError(f"Assignment {index}: no eligible routes found for Any-route search.")
            exhausted: set[str] = set()
            while created < wanted and len(exhausted) < len(routes):
                progressed = False
                for route in routes:
                    if created >= wanted:
                        break
                    if route in exhausted:
                        continue
                    route_created = 0
                    while route_created < 2 and created < wanted:
                        try:
                            _start, _end, _block, matching_rows = _find_one_assignment(
                                sheet,
                                rules,
                                route_override=route,
                            )
                        except ValueError:
                            exhausted.add(route)
                            break
                        _apply_assignment(sheet, next_assignment, matching_rows)
                        results.append(
                            {
                                "guideline": index,
                                "assignment": next_assignment,
                                "rows": matching_rows,
                                "count": len(matching_rows),
                                "route": route,
                            }
                        )
                        next_assignment += 1
                        created += 1
                        route_created += 1
                        progressed = True
                if not progressed:
                    break
            if created == 0:
                raise ValueError(
                    f"Assignment {index}: no blank assignment matched those Any-route rules."
                )
            continue

        # Specific route (batch/bulk): keep searching until count met or no matches remain.
        while created < wanted:
            try:
                _start, _end, _block, matching_rows = _find_one_assignment(sheet, rules)
            except ValueError as exc:
                if created == 0:
                    raise ValueError(f"Assignment {index}: {exc}") from exc
                break
            _apply_assignment(sheet, next_assignment, matching_rows)
            results.append(
                {
                    "guideline": index,
                    "assignment": next_assignment,
                    "rows": matching_rows,
                    "count": len(matching_rows),
                }
            )
            next_assignment += 1
            created += 1

    out = BytesIO()
    workbook.save(out)
    return out.getvalue(), results
