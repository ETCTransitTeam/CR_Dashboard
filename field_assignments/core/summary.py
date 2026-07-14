from __future__ import annotations

from datetime import time
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from field_assignments.core.constants import (
    COL_ASN,
    COL_DIRECTION,
    COL_ROUTE,
    COL_START_LOCATION,
    COL_START_TIME,
)
from field_assignments.core.time_utils import is_blank, normalize_cell, parse_time, time_to_minutes
from field_assignments.core.workbook import read_headers, validate_headers

CoverageGroupBy = str  # "route" | "route_direction" | "route_tod" | "route_direction_tod"


def _coverage_color(pct: float) -> str:
    if pct < 20:
        return "#ffcdd2"
    if pct < 40:
        return "#fff9c4"
    if pct < 70:
        return "#c8e6c9"
    return "#66bb6a"


def _route_sort_key(value: str) -> tuple[int | str, str]:
    return (0, f"{int(value):06d}") if value.isdigit() else (1, value.lower())


def default_tod_ranges() -> list[dict[str, str]]:
    return [{"label": "All Day", "start": "00:00:00", "end": "23:59:59"}]


def _tod_label(start_hm: str, end_hm: str) -> str:
    start = parse_time(start_hm)
    end = parse_time(end_hm)
    if start is None or end is None:
        return f"{start_hm}-{end_hm}"

    def _fmt(t: time) -> str:
        return t.strftime("%I:%M %p").lstrip("0")

    return f"{_fmt(start)}–{_fmt(end)}"


def _bucket_tod(value: object, tod_ranges: list[dict[str, str]]) -> str:
    minutes = time_to_minutes(value)
    if minutes is None:
        return "Unknown"
    # Include seconds for edge cases near midnight by treating time as HH:MM.
    for index, tod in enumerate(tod_ranges):
        start_m = time_to_minutes(tod.get("start"))
        end_m = time_to_minutes(tod.get("end"))
        if start_m is None or end_m is None:
            continue
        # Last bucket may use 23:59:59; treat end of day inclusively.
        if start_m <= end_m:
            if start_m <= minutes <= end_m:
                return tod.get("label") or _tod_label(tod["start"], tod["end"])
        else:
            if minutes >= start_m or minutes <= end_m:
                return tod.get("label") or _tod_label(tod["start"], tod["end"])
    if tod_ranges:
        last = tod_ranges[-1]
        return last.get("label") or _tod_label(last.get("start", ""), last.get("end", ""))
    return "All Day"


def coverage_summary(
    source: bytes,
    sheet_name: str | None = None,
    *,
    group_by: CoverageGroupBy = "route",
    tod_ranges: list[dict[str, str]] | None = None,
) -> pd.DataFrame:
    """
    Build assignment coverage by route, optional direction, and/or time-of-day.

    group_by:
      - route
      - route_direction
      - route_tod
      - route_direction_tod
    """
    workbook = load_workbook(BytesIO(source), data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    headers = read_headers(sheet)
    validate_headers(headers)

    ranges = tod_ranges or default_tod_ranges()
    use_direction = group_by in {"route_direction", "route_direction_tod"}
    use_tod = group_by in {"route_tod", "route_direction_tod"}

    totals: dict[tuple, int] = {}
    assigned: dict[tuple, int] = {}

    for row_number in range(2, sheet.max_row + 1):
        route = normalize_cell(sheet.cell(row=row_number, column=COL_ROUTE).value)
        if not route:
            continue
        key_parts: list[str] = [route]
        if use_direction:
            direction = normalize_cell(sheet.cell(row=row_number, column=COL_DIRECTION).value) or "(none)"
            key_parts.append(direction)
        if use_tod:
            key_parts.append(_bucket_tod(sheet.cell(row=row_number, column=COL_START_TIME).value, ranges))
        key = tuple(key_parts)
        totals[key] = totals.get(key, 0) + 1
        if not is_blank(sheet.cell(row=row_number, column=COL_ASN).value):
            assigned[key] = assigned.get(key, 0) + 1

    rows = []
    for key in sorted(
        totals.keys(),
        key=lambda item: (
            _route_sort_key(item[0]),
            tuple(str(part).lower() for part in item[1:]),
        ),
    ):
        total = totals[key]
        done = assigned.get(key, 0)
        pct = round((done / total) * 100, 1) if total else 0.0
        row: dict[str, object] = {"Route": key[0]}
        col_index = 1
        if use_direction:
            row["Direction"] = key[col_index]
            col_index += 1
        if use_tod:
            row["Time of Day"] = key[col_index]
        row["Total Runs"] = total
        row["Assigned Runs"] = done
        row["% Covered"] = pct
        rows.append(row)

    return pd.DataFrame(rows)


def route_coverage_summary(source: bytes, sheet_name: str | None = None) -> pd.DataFrame:
    return coverage_summary(source, sheet_name, group_by="route")


def style_coverage_dataframe(df: pd.DataFrame) -> pd.io.formats.style.Styler:
    def _row_style(row):
        pct = row["% Covered"]
        color = _coverage_color(float(pct))
        return [f"background-color: {color}"] * len(row)

    return df.style.apply(_row_style, axis=1).format({"% Covered": "{:.1f}%"})


def start_location_routes_summary(source: bytes, sheet_name: str | None = None) -> pd.DataFrame:
    workbook = load_workbook(BytesIO(source), data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    headers = read_headers(sheet)
    validate_headers(headers)

    location_routes: dict[str, set[str]] = {}
    all_routes: set[str] = set()

    for row_number in range(2, sheet.max_row + 1):
        route = normalize_cell(sheet.cell(row=row_number, column=COL_ROUTE).value)
        if not route:
            continue
        all_routes.add(route)
        start_location = normalize_cell(sheet.cell(row=row_number, column=COL_START_LOCATION).value)
        if not start_location:
            continue
        location_routes.setdefault(start_location, set()).add(route)

    total_unique_routes = len(all_routes)
    rows = []
    for start_location in sorted(location_routes.keys(), key=str.lower):
        routes = sorted(location_routes[start_location], key=_route_sort_key)
        route_count = len(routes)
        pct = round((route_count / total_unique_routes) * 100, 1) if total_unique_routes else 0.0
        rows.append(
            {
                "Start Location": start_location,
                "# of Routes": route_count,
                "% of Routes": pct,
                "Routes": ", ".join(routes),
            }
        )

    return pd.DataFrame(rows)


def style_start_location_dataframe(df: pd.DataFrame) -> pd.io.formats.style.Styler:
    return df.style.format({"% of Routes": "{:.1f}%"})
